import json
import sys
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# Manual fills (used only for current_price special coloring)
GREEN  = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
PINK   = PatternFill("solid", fgColor="FFB6C1")

# Gradient CF palette (Excel standard traffic-light)
CF_GREEN  = "63BE7B"
CF_YELLOW = "FFEB84"
CF_RED    = "F8696B"

BOLD      = Font(bold=True)
ITALIC_SM = Font(italic=True, size=9)
SM        = Font(size=9)
INDENT    = "    "


def make_cf_rule(lo, mid, hi, reverse=False):
    """3-color gradient. reverse=True → low value is green (lower is better)."""
    s_color, e_color = (CF_GREEN, CF_RED) if reverse else (CF_RED, CF_GREEN)
    return ColorScaleRule(
        start_type="num", start_value=lo,  start_color=s_color,
        mid_type="num",   mid_value=mid,   mid_color=CF_YELLOW,
        end_type="num",   end_value=hi,    end_color=e_color,
    )


def load_quick(ticker):
    path = f"Outputs/{ticker.upper()}/{ticker.lower()}_quick_metrics.json"
    with open(path) as f:
        return json.load(f)

def load_json(path):
    try:
        with open(path) as f:
            return json.load(f)
    except Exception:
        return {}

def get(d, *keys, default=None):
    for k in keys:
        v = d.get(k)
        if v is not None:
            if isinstance(v, dict):
                vals = [x for x in v.values() if x is not None]
                v = vals[0] if vals else None
            if v is not None:
                return v
    return default


def _calc_rsi(path, period=14):
    all_prices = [v for v in load_json(path).values() if v is not None]
    if len(all_prices) < period + 1:
        return None
    deltas = [all_prices[i] - all_prices[i-1] for i in range(1, len(all_prices))]
    gains  = [max(d, 0)      for d in deltas]
    losses = [abs(min(d, 0)) for d in deltas]
    avg_gain = sum(gains[:period])  / period
    avg_loss = sum(losses[:period]) / period
    for g, l in zip(gains[period:], losses[period:]):
        avg_gain = (avg_gain * (period - 1) + g) / period
        avg_loss = (avg_loss * (period - 1) + l) / period
    if avg_loss == 0:
        return 100.0
    return 100 - (100 / (1 + avg_gain / avg_loss))


def compute_metrics(ticker):
    q = load_quick(ticker)
    t = ticker.lower()
    base = f"Outputs/{ticker.upper()}"

    results = {}

    # Price
    results["current_price"] = get(q, "currentPrice", "regularMarketPrice")
    results["week52_low"]    = get(q, "fiftyTwoWeekLow")
    results["week52_high"]   = get(q, "fiftyTwoWeekHigh")
    results["rsi"]           = _calc_rsi(f"{base}/{t}_price_history.json")

    # Size
    results["market_cap"] = get(q, "marketCap")

    rev = get(q, "totalRevenue")
    if rev is None:
        ttm_is = load_json(f"{base}/{t}_income_statement_ttm.json")
        rev = get(ttm_is, "Total Revenue", "TotalRevenue")
    results["revenue"] = rev

    # Growth
    rev_growth = get(q, "revenueGrowth")
    if rev_growth is None:
        ann = load_json(f"{base}/{t}_income_statement_annual.json")
        tr = ann.get("Total Revenue") or ann.get("TotalRevenue")
        if tr and isinstance(tr, dict):
            vals = [v for v in tr.values() if v is not None]
            if len(vals) >= 2:
                rev_growth = (vals[0] - vals[1]) / abs(vals[1]) if vals[1] else None
    results["rev_growth"] = rev_growth

    # Margins
    gm = get(q, "grossMargins")
    if gm is None:
        ttm_is = load_json(f"{base}/{t}_income_statement_ttm.json")
        total_rev = get(ttm_is, "Total Revenue", "TotalRevenue")
        cost_rev  = get(ttm_is, "Cost Of Revenue", "CostOfRevenue", "Cost of Revenue")
        if total_rev and cost_rev is not None:
            gm = (total_rev - cost_rev) / total_rev
    results["gross_margin"] = gm

    op_margin = get(q, "operatingMargins")
    if op_margin is None:
        ttm_is    = load_json(f"{base}/{t}_income_statement_ttm.json")
        total_rev = get(ttm_is, "Total Revenue", "TotalRevenue") or rev
        op_inc    = get(ttm_is, "Operating Income", "OperatingIncome", "Total Operating Income As Reported")
        if total_rev and op_inc is not None:
            op_margin = op_inc / total_rev
    results["op_margin"] = op_margin

    ni_margin = get(q, "profitMargins")
    if ni_margin is None:
        ttm_is    = load_json(f"{base}/{t}_income_statement_ttm.json")
        total_rev = get(ttm_is, "Total Revenue", "TotalRevenue") or rev
        ni        = get(ttm_is, "Net Income", "NetIncome", "Net Income Common Stockholders")
        if total_rev and ni is not None:
            ni_margin = ni / total_rev
    results["ni_margin"] = ni_margin

    # Returns / leverage
    results["roe"] = get(q, "returnOnEquity")

    de_raw = get(q, "debtToEquity")
    results["de"] = de_raw / 100 if de_raw is not None else None

    ttm_is      = load_json(f"{base}/{t}_income_statement_ttm.json")
    ebit        = get(ttm_is, "Operating Income", "OperatingIncome", "Total Operating Income As Reported", "EBIT")
    interest_exp = get(ttm_is, "Interest Expense", "InterestExpense")
    results["interest_cov"] = (ebit / abs(interest_exp)
                               if ebit is not None and interest_exp and interest_exp != 0
                               else None)

    cur_ratio = get(q, "currentRatio")
    if cur_ratio is None:
        bs        = load_json(f"{base}/{t}_balance_sheet_quarterly.json")
        cur_assets = get(bs, "Current Assets", "CurrentAssets", "Total Current Assets")
        cur_liab   = get(bs, "Current Liabilities", "CurrentLiabilities",
                         "Total Current Liabilities Net Minority Interest")
        if cur_assets and cur_liab:
            cur_ratio = cur_assets / cur_liab
    results["cur_ratio"] = cur_ratio

    fcf = get(q, "freeCashflow")
    results["fcf_margin"] = (fcf / rev) if (fcf is not None and rev) else None
    results["r40"] = ((rev_growth * 100 + op_margin * 100)
                      if (rev_growth is not None and op_margin is not None) else None)

    # Valuation
    results["trailing_pe"]   = get(q, "trailingPE")
    results["forward_pe"]    = get(q, "forwardPE")
    results["peg"]           = get(q, "pegRatio", "trailingPegRatio")
    results["price_to_sales"] = get(q, "priceToSalesTrailing12Months")

    results["payout_ratio"] = get(q, "payoutRatio")
    results["sector"]       = q.get("sector", "")

    forward_dy  = q.get("dividendYield")
    trailing_dy = q.get("trailingAnnualDividendYield")
    if forward_dy:
        results["dividend_yield"] = forward_dy / 100
    elif trailing_dy:
        results["dividend_yield"] = trailing_dy
    else:
        results["dividend_yield"] = None

    return results


# ── Metric definitions ───────────────────────────────────────────────────────
# num_fmt : Excel number-format string applied to the cell
# scale   : multiply raw value by this before writing to cell (default 1)
# cf      : (lo, mid, hi, reverse) for ColorScaleRule, or None

METRICS = [
    {
        "key":     "current_price",
        "label":   "1. Current Price",
        "desc":    "Most recent market price of the stock.",
        "bench":   "Green = bottom third of 52-wk range | Yellow = middle third | Pink = top third",
        "num_fmt": '$#,##0.00',
        "scale":   1,
        "cf":      None,   # handled via manual fill
    },
    {
        "key":     "week52_low",
        "label":   "2. 52-Week Low",
        "desc":    "Lowest traded price over the trailing 52 weeks.",
        "bench":   "Informational — no threshold coloring",
        "num_fmt": '$#,##0.00',
        "scale":   1,
        "cf":      None,
    },
    {
        "key":     "week52_high",
        "label":   "3. 52-Week High",
        "desc":    "Highest traded price over the trailing 52 weeks.",
        "bench":   "Informational — no threshold coloring",
        "num_fmt": '$#,##0.00',
        "scale":   1,
        "cf":      None,
    },
    {
        "key":     "rsi",
        "label":   "4. RSI (14-Day)",
        "desc":    "Wilder's 14-day Relative Strength Index. Measures momentum and overbought/oversold conditions.",
        "bench":   "≤30 → Oversold / potential buy | 30–70 → Neutral | ≥70 → Overbought / potential sell",
        "num_fmt": '0.0',
        "scale":   1,
        "cf":      (20, 50, 80, True),   # lower = oversold = green
    },
    {
        "key":     "market_cap",
        "label":   "5. Market Cap",
        "desc":    "Total market value of all outstanding shares (Price × Shares Outstanding).",
        "bench":   "Informational — no threshold coloring",
        "num_fmt": '$#,##0.0"B"',
        "scale":   1e-9,
        "cf":      None,
    },
    {
        "key":     "revenue",
        "label":   "6. Revenue (TTM)",
        "desc":    "Total revenue over the trailing twelve months.",
        "bench":   "Informational — no threshold coloring",
        "num_fmt": '$#,##0.0"B"',
        "scale":   1e-9,
        "cf":      None,
    },
    {
        "key":     "rev_growth",
        "label":   "7. Revenue Growth Rate (YoY)",
        "desc":    "YoY revenue growth from quick metrics or computed from two most recent annual periods.",
        "bench":   ">20% → Strong | 10–20% → Solid | <10% → Slow",
        "num_fmt": '0.0%',
        "scale":   1,
        "cf":      (-0.05, 0.10, 0.25, False),
    },
    {
        "key":     "gross_margin",
        "label":   "8. Gross Margin",
        "desc":    "(Total Revenue − Cost of Revenue) / Total Revenue (TTM).",
        "bench":   ">60% → High quality | 40–60% → Decent | <40% → Watch for pricing pressure",
        "num_fmt": '0.0%',
        "scale":   1,
        "cf":      (0.15, 0.40, 0.65, False),
    },
    {
        "key":     "op_margin",
        "label":   "9. Operating Margin",
        "desc":    "Operating Income (TTM) / Total Revenue (TTM).",
        "bench":   ">30% → Strong pricing power | 15–30% → Decent | <15% → Watch for cost pressure",
        "num_fmt": '0.0%',
        "scale":   1,
        "cf":      (0.0, 0.15, 0.35, False),
    },
    {
        "key":     "ni_margin",
        "label":   "10. Net Income Margin",
        "desc":    "Net Income (TTM) / Total Revenue (TTM). Bottom-line profitability after all expenses.",
        "bench":   ">20% → Strong | 10–20% → Decent | <10% → Thin",
        "num_fmt": '0.0%',
        "scale":   1,
        "cf":      (0.0, 0.10, 0.25, False),
    },
    {
        "key":     "roe",
        "label":   "11. Return on Equity (ROE)",
        "desc":    "Net Income (TTM) / Stockholders Equity (MRQ).",
        "bench":   "≥20% → Ideal | ≥15% → Good | <15% → Below threshold",
        "num_fmt": '0.0%',
        "scale":   1,
        "cf":      (0.0, 0.12, 0.25, False),
    },
    {
        "key":     "de",
        "label":   "12. Debt-to-Equity (D/E)",
        "desc":    "Total Debt (MRQ) / Stockholders Equity (MRQ). Measures financial leverage.",
        "bench":   "<0.5 → Very conservative | 0.5–1.0 → Healthy | 1.0–2.0 → Moderate | >2.0 → High risk",
        "num_fmt": '0%',
        "scale":   1,
        "cf":      (0.0, 1.0, 3.0, True),  # lower is better
    },
    {
        "key":     "interest_cov",
        "label":   "13. Interest Coverage",
        "desc":    "EBIT (TTM) / Interest Expense (TTM). Ability to service debt from operating earnings.",
        "bench":   ">10× → Very safe | 5–10× → Adequate | 3–5× → Watch | <3× → At risk",
        "num_fmt": '0.0"×"',
        "scale":   1,
        "cf":      (1.0, 5.0, 15.0, False),
    },
    {
        "key":     "cur_ratio",
        "label":   "14. Current Ratio",
        "desc":    "Current Assets (MRQ) / Current Liabilities (MRQ). Short-term liquidity.",
        "bench":   ">2.0 → Very liquid | 1.5–2.0 → Healthy | 1.0–1.5 → Adequate | <1.0 → Liquidity risk",
        "num_fmt": '0.00',
        "scale":   1,
        "cf":      (0.5, 1.5, 3.0, False),
    },
    {
        "key":     "fcf_margin",
        "label":   "15. FCF Margin",
        "desc":    "Free Cash Flow (TTM) / Total Revenue (TTM). Cash generated per dollar of revenue after capex.",
        "bench":   ">20% → High quality | 10–20% → Solid | <10% → Low",
        "num_fmt": '0.0%',
        "scale":   1,
        "cf":      (-0.05, 0.10, 0.30, False),
    },
    {
        "key":     "r40",
        "label":   "16. Rule of 40",
        "desc":    "Revenue Growth Rate (YoY%) + Operating Margin (TTM%). Balances growth and profitability.",
        "bench":   ">40 → Healthy/investible | 30–40 → Borderline | <30 → Warning zone",
        "num_fmt": '0.0',
        "scale":   1,
        "cf":      (0, 25, 50, False),
    },
    {
        "key":     "trailing_pe",
        "label":   "17a. Trailing P/E",
        "desc":    "Price / Trailing Twelve Months EPS. How much investors pay per dollar of past earnings.",
        "bench":   "<15 → Cheap | 15–25 → Fair | >25 → Expensive",
        "num_fmt": '0.0',
        "scale":   1,
        "cf":      (10, 25, 50, True),  # lower is better
    },
    {
        "key":     "forward_pe",
        "label":   "17b. Forward P/E",
        "desc":    "Price / Next Twelve Months EPS Estimate. Investors' view of future earnings power.",
        "bench":   "<15 → Cheap | 15–25 → Fair | >25 → Expensive",
        "num_fmt": '0.0',
        "scale":   1,
        "cf":      (8, 20, 40, True),   # lower is better
    },
    {
        "key":     "peg",
        "label":   "17c. PEG Ratio",
        "desc":    "Trailing P/E / Earnings Growth Rate. Adjusts valuation for expected growth.",
        "bench":   "<1 → Potentially undervalued | 1–2 → Fair | >2 → Expensive relative to growth",
        "num_fmt": '0.00',
        "scale":   1,
        "cf":      (0.5, 1.5, 3.0, True),  # lower is better
    },
    {
        "key":     "price_to_sales",
        "label":   "17d. Price / Sales (TTM)",
        "desc":    "Market Cap / Total Revenue (TTM). Useful for low-earnings companies; lower is cheaper.",
        "bench":   "<3 → Cheap | 3–6 → Fair | >6 → Expensive",
        "num_fmt": '0.0',
        "scale":   1,
        "cf":      (1, 4, 10, True),    # lower is better
    },
    {
        "key":     "dividend_yield",
        "label":   "18. Dividend Yield",
        "desc":    "Annual dividend per share / current price. Shown when available; N/A otherwise.",
        "bench":   ">4% → High yield | 2–4% → Moderate | <2% → Low yield",
        "num_fmt": '0.0%',
        "scale":   1,
        "cf":      (0.005, 0.025, 0.05, False),  # higher is better
    },
    {
        "key":     "payout_ratio",
        "label":   "19. Dividend Payout Ratio",
        "desc":    "Dividends paid / Net Income (TTM). Measures what fraction of earnings is returned to shareholders as dividends. Note: for REITs, net income understates true cash earnings — AFFO payout ratio (dividends / AFFO) is the appropriate metric; values shown here for REITs will appear artificially high and should be interpreted with caution.",
        "bench":   "<50% → Sustainable | 50–75% → Moderate | >75% → High / potential cut risk | >100% → Unsustainable (REITs: use AFFO payout ratio instead)",
        "num_fmt": '0%',
        "scale":   1,
        "cf":      (0.2, 0.6, 1.0, True),  # lower is better (more sustainable)
    },
]


# ── Helpers ──────────────────────────────────────────────────────────────────

def color_current_price(metrics):
    """Manual fill: green = bottom third, yellow = middle third, pink = top third of 52-wk range."""
    price = metrics.get("current_price")
    low   = metrics.get("week52_low")
    high  = metrics.get("week52_high")
    if price is None or low is None or high is None or high == low:
        return None
    pos = (price - low) / (high - low)
    if pos < 1/3:   return GREEN
    if pos <= 2/3:  return YELLOW
    return PINK


def _short_comment(key, val, metrics=None):
    if val is None:
        return ""
    if key == "current_price":
        low  = (metrics or {}).get("week52_low")
        high = (metrics or {}).get("week52_high")
        if low is not None and high is not None and high != low:
            pos  = (val - low) / (high - low)
            zone = ("Near 52-wk low" if pos < 1/3
                    else ("Mid-range" if pos <= 2/3 else "Near 52-wk high"))
            return f"{zone} ({pos*100:.0f}% of range)"
        return ""
    if key == "rsi":
        if val >= 70: return "Overbought (≥70)"
        if val <= 30: return "Oversold (≤30)"
        return "Neutral (30–70)"
    if key == "rev_growth":
        if val > 0.20:  return "Strong (>20%)"
        if val >= 0.10: return "Solid (10–20%)"
        return "Slow (<10%)"
    if key == "gross_margin":
        if val > 0.60:  return "High quality (>60%)"
        if val >= 0.40: return "Decent (40–60%)"
        return "Watch for pricing pressure (<40%)"
    if key == "op_margin":
        if val > 0.30:  return "Strong pricing power (>30%)"
        if val >= 0.15: return "Decent (15–30%)"
        return "Watch for cost pressure (<15%)"
    if key == "ni_margin":
        if val > 0.20:  return "Strong (>20%)"
        if val >= 0.10: return "Decent (10–20%)"
        return "Thin (<10%)"
    if key == "roe":
        if val >= 0.20: return "Ideal (≥20%)"
        if val >= 0.15: return "Good (≥15%)"
        return "Below threshold (<15%)"
    if key == "de":
        if val < 0.5:  return "Very conservative"
        if val <= 1.0: return "Healthy"
        if val <= 2.0: return "Moderate leverage"
        return "High risk"
    if key == "interest_cov":
        if val > 10: return "Very safe (>10×)"
        if val >= 5: return "Adequate (5–10×)"
        if val >= 3: return "Watch (3–5×)"
        return "At risk (<3×)"
    if key == "cur_ratio":
        if val > 2.0:  return "Very liquid (>2.0)"
        if val >= 1.5: return "Healthy (1.5–2.0)"
        if val >= 1.0: return "Adequate (1.0–1.5)"
        return "Liquidity risk (<1.0)"
    if key == "fcf_margin":
        if val > 0.20:  return "High quality (>20%)"
        if val >= 0.10: return "Solid (10–20%)"
        return "Low (<10%)"
    if key == "r40":
        if val > 40:  return "Healthy / investible (>40)"
        if val >= 30: return "Borderline (30–40)"
        return "Warning zone (<30)"
    if key == "peg":
        if val < 1:  return "Potentially undervalued (<1)"
        if val <= 2: return "Fair (1–2)"
        return "Expensive relative to growth (>2)"
    if key == "payout_ratio":
        sector = (metrics or {}).get("sector", "")
        if sector == "Real Estate":
            return "⚠ REIT — use AFFO payout ratio; net income understates cash earnings"
        if val < 0.5:  return "Sustainable (<50%)"
        if val <= 0.75: return "Moderate (50–75%)"
        if val <= 1.0:  return "High — potential cut risk (75–100%)"
        return "Unsustainable (>100%)"
    return ""


# ── Sheet writers ─────────────────────────────────────────────────────────────

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF")


def _write_value_cell(ws, row, col, val, m, ticker_metrics=None):
    """Write a numeric value cell with number format, manual fill or CF rule."""
    cell = ws.cell(row=row, column=col)
    cell.alignment = Alignment(horizontal="center")

    if val is None:
        cell.value = "N/A"
        return

    cell.value = val * m["scale"]
    cell.number_format = m["num_fmt"]

    # Current price: manual fill based on 52-wk range
    if m["key"] == "current_price" and ticker_metrics is not None:
        fill = color_current_price(ticker_metrics)
        if fill:
            cell.fill = fill


def _apply_cf(ws, cell_range, m):
    """Apply ColorScaleRule to a cell range for metrics that have CF defined."""
    if m["cf"] is None or m["key"] == "current_price":
        return
    lo, mid, hi, reverse = m["cf"]
    rule = make_cf_rule(lo * m["scale"], mid * m["scale"], hi * m["scale"], reverse)
    ws.conditional_formatting.add(cell_range, rule)


def write_ticker_sheet(wb, ticker, metrics):
    ws = wb.create_sheet(title=ticker)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 42

    ws.cell(row=1, column=1, value=f"{ticker} — Key Stock Metrics").font = Font(bold=True, size=13)
    ws.append([])

    for col, heading in enumerate(["Metric", "Value", "Comment"], start=1):
        c = ws.cell(row=3, column=col, value=heading)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    for i, m in enumerate(METRICS):
        row = i + 4
        val = metrics.get(m["key"])
        ws.cell(row=row, column=1, value=m["label"]).font = BOLD
        _write_value_cell(ws, row, 2, val, m, ticker_metrics=metrics)
        ws.cell(row=row, column=3, value=_short_comment(m["key"], val, metrics)).font = SM
        cell_addr = f"B{row}"
        _apply_cf(ws, cell_addr, m)


def write_comparison_sheet(wb, tickers, all_metrics):
    ws = wb.create_sheet(title="Comparison")
    wb.move_sheet(ws, offset=-(len(wb.sheetnames) - 1))

    ws.column_dimensions["A"].width = 28
    for col in range(2, len(tickers) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # ── Header row ────────────────────────────────────────────────────────────
    c = ws.cell(row=1, column=1, value="Metric")
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = Alignment(horizontal="center")
    for col, ticker in enumerate(tickers, start=2):
        c = ws.cell(row=1, column=col, value=ticker)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = Alignment(horizontal="center")

    # ── Data rows ─────────────────────────────────────────────────────────────
    for r, m in enumerate(METRICS, start=2):
        ws.cell(row=r, column=1, value=m["label"]).font = BOLD
        for col, ticker in enumerate(tickers, start=2):
            val = all_metrics[ticker].get(m["key"])
            _write_value_cell(ws, r, col, val, m,
                              ticker_metrics=all_metrics[ticker] if m["key"] == "current_price" else None)

        # Apply CF across all ticker columns for this row
        if len(tickers) == 1:
            cf_range = f"{get_column_letter(2)}{r}"
        else:
            cf_range = f"{get_column_letter(2)}{r}:{get_column_letter(len(tickers)+1)}{r}"
        _apply_cf(ws, cf_range, m)

    # ── Metric descriptions & benchmarks ─────────────────────────────────────
    desc_start = len(METRICS) + 3
    ws.cell(row=desc_start, column=1,
            value="Metric Descriptions & Benchmarks").font = Font(bold=True, size=12)
    row = desc_start + 2
    for m in METRICS:
        ws.cell(row=row,   column=1, value=m["label"]).font = BOLD
        ws.cell(row=row+1, column=1, value=f"{INDENT}Description:").font = ITALIC_SM
        ws.cell(row=row+1, column=2, value=m["desc"]).font = SM
        ws.cell(row=row+2, column=1, value=f"{INDENT}Benchmarks:").font = ITALIC_SM
        ws.cell(row=row+2, column=2, value=m["bench"]).font = SM
        row += 4


# ── Entry point ───────────────────────────────────────────────────────────────

def main(tickers):
    all_metrics = {}
    for t in tickers:
        print(f"Computing metrics for {t}...")
        all_metrics[t] = compute_metrics(t)

    wb = Workbook()
    wb.remove(wb.active)

    for t in tickers:
        write_ticker_sheet(wb, t, all_metrics[t])

    write_comparison_sheet(wb, tickers, all_metrics)

    out = f"Outputs/key_stock_metrics_{date.today().strftime('%Y%m%d')}.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    return out


if __name__ == "__main__":
    if len(sys.argv) > 1:
        tickers = [t.upper() for t in sys.argv[1:]]
    else:
        sys.path.insert(0, ".")
        from yahoo_finance_data import load_tickers
        tickers = load_tickers()
    main(tickers)
